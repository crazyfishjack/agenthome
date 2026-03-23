"""
生成分类目录索引脚本
用于根据分类中的文件，生成或更新目录索引
支持PDF、Word、Excel、TXT、Markdown格式文档
"""

import sys
import json
import argparse
from pathlib import Path
from datetime import datetime
import os


# 支持的文档类型映射
DOC_TYPE_MAP = {
    '.pdf': 'pdf',
    '.doc': 'doc',
    '.docx': 'docx',
    '.xls': 'xls',
    '.xlsx': 'xlsx',
    '.txt': 'txt',
    '.md': 'md'
}


def get_file_update_time(file_path):
    """获取文件的修改时间"""
    timestamp = os.path.getmtime(file_path)
    return datetime.fromtimestamp(timestamp).strftime("%Y-%m-%d %H:%M:%S")


def extract_text_from_txt(file_path):
    """从TXT文件中提取文本内容"""
    try:
        # 尝试多种编码
        encodings = ['utf-8', 'gbk', 'gb2312', 'utf-16']
        for encoding in encodings:
            try:
                with open(file_path, 'r', encoding=encoding) as f:
                    return f.read()
            except UnicodeDecodeError:
                continue
        return ""
    except Exception as e:
        print(f"读取TXT文件失败: {str(e)}", file=sys.stderr)
        return ""


def extract_text_from_pdf(file_path):
    """从PDF文件中提取文本内容"""
    try:
        import PyPDF2
        text = ""
        with open(file_path, 'rb') as file:
            reader = PyPDF2.PdfReader(file)
            for page in reader.pages:
                text += page.extract_text() + "\n"
        return text
    except ImportError:
        print("错误: 需要安装PyPDF2库来处理PDF文件", file=sys.stderr)
        return ""
    except Exception as e:
        print(f"读取PDF文件失败: {str(e)}", file=sys.stderr)
        return ""


def extract_text_from_word(file_path):
    """从Word文件中提取文本内容"""
    try:
        from docx import Document
        doc = Document(file_path)
        text = ""
        for paragraph in doc.paragraphs:
            text += paragraph.text + "\n"
        return text
    except ImportError:
        print("错误: 需要安装python-docx库来处理Word文件", file=sys.stderr)
        return ""
    except Exception as e:
        print(f"读取Word文件失败: {str(e)}", file=sys.stderr)
        return ""


def extract_text_from_excel(file_path):
    """从Excel文件中提取文本内容"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        text = ""
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            text += f"=== {sheet_name} ===\n"
            for row in sheet.iter_rows(values_only=True):
                row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                if row_text.strip():
                    text += row_text + "\n"
        return text
    except ImportError:
        print("错误: 需要安装openpyxl库来处理Excel文件", file=sys.stderr)
        return ""
    except Exception as e:
        print(f"读取Excel文件失败: {str(e)}", file=sys.stderr)
        return ""


def extract_text_from_markdown(file_path):
    """从Markdown文件中提取文本内容"""
    try:
        # Markdown文件使用utf-8编码读取
        with open(file_path, 'r', encoding='utf-8') as f:
            return f.read()
    except Exception as e:
        print(f"读取Markdown文件失败: {str(e)}", file=sys.stderr)
        return ""


def extract_text_from_document(file_path, doc_type):
    """根据文档类型提取文本内容"""
    if doc_type == 'txt':
        return extract_text_from_txt(file_path)
    elif doc_type == 'pdf':
        return extract_text_from_pdf(file_path)
    elif doc_type in ['doc', 'docx']:
        return extract_text_from_word(file_path)
    elif doc_type in ['xls', 'xlsx']:
        return extract_text_from_excel(file_path)
    elif doc_type == 'md':
        return extract_text_from_markdown(file_path)
    else:
        return ""


def generate_index(category, knowledge_base_path):
    """生成分类目录索引"""
    category_path = Path(knowledge_base_path) / category

    # 检查分类目录是否存在
    if not category_path.exists():
        return {
            "success": False,
            "message": f"分类目录不存在: {category}",
            "category": category,
            "document_count": 0,
            "update_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }

    # 获取目录下所有支持的文档文件
    documents = []
    supported_extensions = list(DOC_TYPE_MAP.keys())

    for file_path in category_path.iterdir():
        if file_path.is_file() and file_path.suffix.lower() in supported_extensions:
            doc_type = DOC_TYPE_MAP[file_path.suffix.lower()]
            update_time = get_file_update_time(file_path)

            # 计算相对于知识库根目录的路径
            relative_path = file_path.relative_to(Path(knowledge_base_path).parent)

            doc_info = {
                "doc_name": file_path.name,
                "doc_path": str(relative_path).replace("\\", "/"),
                "doc_type": doc_type,
                "update_time": update_time,
                "summary": "在此输入摘要"
            }
            documents.append(doc_info)

    # 按文档名称排序
    documents.sort(key=lambda x: x["doc_name"])

    # 生成索引数据
    index_data = {
        "category_name": category,
        "update_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "documents": documents
    }

    # 保存索引文件
    index_file_path = category_path / "index.json"
    try:
        with open(index_file_path, 'w', encoding='utf-8') as f:
            json.dump(index_data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        return {
            "success": False,
            "message": f"保存索引文件失败: {str(e)}",
            "category": category,
            "document_count": len(documents),
            "update_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }

    return {
        "success": True,
        "message": "索引生成成功",
        "category": category,
        "document_count": len(documents),
        "update_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }


def get_knowledge_base_path():
    """获取知识库根目录路径"""
    # 获取项目根目录（从scripts目录向上5级）
    project_root = Path(__file__).parent.parent.parent.parent.parent
    knowledge_base_path = project_root / "rag" / "rag_document_lib"
    return str(knowledge_base_path)


def main():
    parser = argparse.ArgumentParser(description='生成分类目录索引')
    parser.add_argument('--category', type=str, required=True, help='分类名称')
    args = parser.parse_args()

    category = args.category
    knowledge_base_path = get_knowledge_base_path()

    # 生成索引
    result = generate_index(category, knowledge_base_path)

    # 输出JSON格式结果
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
