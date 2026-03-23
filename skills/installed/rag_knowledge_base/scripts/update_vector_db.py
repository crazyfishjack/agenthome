"""
更新向量数据库脚本
用于将文档分块并存储到向量数据库
支持按分类更新或按文档更新
支持PDF、Word、Excel、TXT、Markdown格式文档
"""

import sys
import json
import argparse
import os
import re
from pathlib import Path
from datetime import datetime
import hashlib

# 尝试加载 .env 文件（从项目根目录）
try:
    from dotenv import load_dotenv
    project_root = Path(__file__).parent.parent.parent.parent.parent
    env_path = project_root / ".env"
    if env_path.exists():
        load_dotenv(dotenv_path=env_path)
except ImportError:
    pass

# 文档处理库
try:
    import PyPDF2
except ImportError:
    PyPDF2 = None

try:
    from docx import Document
except ImportError:
    Document = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

# 向量数据库
try:
    import chromadb
    from chromadb.config import Settings
except ImportError:
    chromadb = None

# HTTP请求
try:
    import httpx
except ImportError:
    httpx = None


# 配置
EMBEDDING_API_KEY = os.environ.get("EMBEDDING_API_KEY", "")
EMBEDDING_API_URL = "https://dashscope.aliyuncs.com/compatible-mode/v1/embeddings"
EMBEDDING_MODEL = "text-embedding-v4"

# 路径配置
PROJECT_ROOT = Path(__file__).parent.parent.parent.parent.parent
RAG_LIB_PATH = PROJECT_ROOT / "rag" / "rag_document_lib"
CHROMA_DB_PATH = PROJECT_ROOT / "rag" / "chroma_db"


class DocumentCleaner:
    """文档清洗类"""

    @staticmethod
    def clean_text(text: str) -> str:
        """
        清洗文本内容
        只去除多余空行和空白字符,保留所有有用信息
        """
        if not text:
            return ""

        # 1. 统一换行符
        text = text.replace('\r\n', '\n').replace('\r', '\n')

        # 2. 去除多余空行（连续超过2个空行）
        text = re.sub(r'\n\s*\n\s*\n+', '\n\n', text)

        # 3. 去除行首行尾多余空格
        lines = text.split('\n')
        lines = [line.strip() for line in lines]
        text = '\n'.join(lines)

        # 4. 去除空白行
        lines = [line for line in text.split('\n') if line.strip()]
        text = '\n'.join(lines)

        return text.strip()


class DocumentExtractor:
    """文档提取类"""

    @staticmethod
    def extract_text(file_path: Path) -> str:
        """
        从文档中提取文本
        支持PDF、Word、Excel、TXT、Markdown格式
        """
        file_ext = file_path.suffix.lower()

        if file_ext == '.pdf':
            return DocumentExtractor._extract_pdf(file_path)
        elif file_ext in ['.doc', '.docx']:
            return DocumentExtractor._extract_word(file_path)
        elif file_ext in ['.xls', '.xlsx']:
            return DocumentExtractor._extract_excel(file_path)
        elif file_ext == '.txt':
            return DocumentExtractor._extract_txt(file_path)
        elif file_ext == '.md':
            return DocumentExtractor._extract_markdown(file_path)
        else:
            raise ValueError(f"不支持的文件格式: {file_ext}")

    @staticmethod
    def _extract_pdf(file_path: Path) -> str:
        """提取PDF文本"""
        if PyPDF2 is None:
            raise ImportError("需要安装 PyPDF2 库: pip install PyPDF2")

        text_parts = []
        try:
            with open(file_path, 'rb') as f:
                pdf_reader = PyPDF2.PdfReader(f)
                for page in pdf_reader.pages:
                    try:
                        text = page.extract_text()
                        if text:
                            text_parts.append(text)
                    except Exception as e:
                        print(f"警告: 提取PDF页面时出错: {e}")
                        continue
        except Exception as e:
            raise Exception(f"读取PDF文件失败: {e}")

        return '\n'.join(text_parts)

    @staticmethod
    def _extract_word(file_path: Path) -> str:
        """提取Word文本"""
        if Document is None:
            raise ImportError("需要安装 python-docx 库: pip install python-docx")

        try:
            doc = Document(file_path)
            text_parts = []
            for paragraph in doc.paragraphs:
                if paragraph.text.strip():
                    text_parts.append(paragraph.text.strip())
            return '\n'.join(text_parts)
        except Exception as e:
            raise Exception(f"读取Word文件失败: {e}")

    @staticmethod
    def _extract_excel(file_path: Path) -> str:
        """提取Excel文本"""
        if openpyxl is None:
            raise ImportError("需要安装 openpyxl 库: pip install openpyxl")

        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            text_parts = []

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_text = []

                for row in sheet.iter_rows(values_only=True):
                    row_text = []
                    for cell in row:
                        if cell is not None and str(cell).strip():
                            row_text.append(str(cell).strip())
                    if row_text:
                        sheet_text.append(' | '.join(row_text))

                if sheet_text:
                    text_parts.append(f"【{sheet_name}】")
                    text_parts.extend(sheet_text)

            workbook.close()
            return '\n'.join(text_parts)
        except Exception as e:
            raise Exception(f"读取Excel文件失败: {e}")

    @staticmethod
    def _extract_txt(file_path: Path) -> str:
        """
        提取TXT文本
        优先使用UTF-8编码，遇到错误字符时跳过并尝试下一个编码
        保持向后兼容，支持多种编码作为备选
        """
        # 优先使用UTF-8编码，其他编码作为备选
        encodings = ['utf-8', 'gbk', 'gb2312', 'utf-16', 'latin-1']

        last_error = None

        for encoding in encodings:
            try:
                # 尝试使用严格模式读取
                with open(file_path, 'r', encoding=encoding) as f:
                    text = f.read()
                    if text:  # 确保读取到内容
                        return text
            except UnicodeDecodeError as e:
                # 记录错误，继续尝试下一个编码
                last_error = f"{encoding}: {str(e)}"
                continue
            except Exception as e:
                # 其他异常也记录，但继续尝试
                last_error = f"{encoding}: {str(e)}"
                continue

        # 如果所有编码都失败，尝试使用errors='ignore'参数跳过非法字符
        for encoding in encodings:
            try:
                with open(file_path, 'r', encoding=encoding, errors='ignore') as f:
                    text = f.read()
                    if text:  # 确保读取到内容
                        print(f"警告: 使用 {encoding} 编码读取文件时跳过了非法字符")
                        return text
            except Exception as e:
                last_error = f"{encoding} (ignore): {str(e)}"
                continue

        # 所有尝试都失败
        raise Exception(f"无法解码TXT文件，尝试了多种编码: {encodings}。最后错误: {last_error}")

    @staticmethod
    def _extract_markdown(file_path: Path) -> str:
        """提取Markdown文本"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                return f.read()
        except Exception as e:
            raise Exception(f"读取Markdown文件失败: {e}")


class TextChunker:
    """文本分块类"""

    @staticmethod
    def chunk_text(text: str, chunk_size: int = 300, chunk_overlap: int = 40) -> list:
        """
        将文本分块
        """
        if not text:
            return []

        chunks = []
        start = 0
        text_length = len(text)

        while start < text_length:
            end = start + chunk_size

            # 如果不是最后一块，尝试在句号、问号、感叹号处切分
            if end < text_length:
                # 查找最近的句子结束符
                for i in range(end, max(start + chunk_size // 2, start), -1):
                    if text[i] in ['。', '！', '？', '.', '!', '?', '\n']:
                        end = i + 1
                        break

            chunk = text[start:end].strip()
            if chunk:
                chunks.append(chunk)

            start = end - chunk_overlap

        return chunks


class EmbeddingGenerator:
    """Embedding生成类"""

    @staticmethod
    def generate_embeddings(texts: list) -> list:
        """
        生成文本的embedding向量
        """
        if httpx is None:
            raise ImportError("需要安装 httpx 库: pip install httpx")

        headers = {
            'Authorization': f'Bearer {EMBEDDING_API_KEY}',
            'Content-Type': 'application/json'
        }

        embeddings = []

        # 批量处理，每次最多处理10个文本（阿里云DashScope限制）
        batch_size = 10
        for i in range(0, len(texts), batch_size):
            batch = texts[i:i + batch_size]

            data = {
                'model': EMBEDDING_MODEL,
                'input': batch,
                'encoding_format': 'float'
            }

            try:
                response = httpx.post(
                    EMBEDDING_API_URL,
                    headers=headers,
                    json=data,
                    timeout=30.0
                )
                response.raise_for_status()

                result = response.json()
                batch_embeddings = [item['embedding'] for item in result['data']]
                embeddings.extend(batch_embeddings)

            except Exception as e:
                raise Exception(f"生成embedding失败: {e}")

        return embeddings


class VectorDatabase:
    """向量数据库管理类"""

    def __init__(self, db_path: Path):
        if chromadb is None:
            raise ImportError("需要安装 chromadb 库: pip install chromadb")

        self.client = chromadb.PersistentClient(path=str(db_path))
        self.collection = self.client.get_or_create_collection(
            name="rag_knowledge_base",
            metadata={"hnsw:space": "cosine"}
        )

    def delete_by_category(self, category: str) -> int:
        """
        删除指定分类的所有分块
        """
        try:
            # 查询该分类的所有分块ID
            results = self.collection.get(
                where={"category": category}
            )

            if results['ids']:
                self.collection.delete(ids=results['ids'])
                return len(results['ids'])
            return 0
        except Exception as e:
            raise Exception(f"删除分类分块失败: {e}")

    def delete_by_document(self, category: str, doc_name: str) -> int:
        """
        删除指定文档的所有分块
        """
        try:
            # Chroma不支持多条件AND查询，先按category查询
            results = self.collection.get(
                where={"category": category}
            )

            # 过滤出指定文档的分块ID
            ids_to_delete = []
            for idx, metadata in enumerate(results['metadatas']):
                if metadata.get('doc_name') == doc_name:
                    ids_to_delete.append(results['ids'][idx])

            if ids_to_delete:
                self.collection.delete(ids=ids_to_delete)
                return len(ids_to_delete)
            return 0
        except Exception as e:
            raise Exception(f"删除文档分块失败: {e}")

    def add_chunks(self, chunks_data: list) -> int:
        """
        添加分块到向量数据库
        chunks_data: 包含分块信息的字典列表
        """
        try:
            ids = []
            embeddings = []
            metadatas = []
            documents = []

            for chunk_data in chunks_data:
                ids.append(chunk_data['chunk_id'])
                embeddings.append(chunk_data['embedding'])
                metadatas.append({
                    "chunk_id": chunk_data['chunk_id'],
                    "category": chunk_data['category'],
                    "doc_name": chunk_data['doc_name'],
                    "doc_update_time": chunk_data['doc_update_time'],
                    "chunk_content": chunk_data['chunk_content'],
                    "chunk_token_size": chunk_data['chunk_token_size'],
                    "chunk_index": chunk_data['chunk_index'],
                    "doc_path": chunk_data['doc_path'],
                    "create_time": chunk_data['create_time']
                })
                documents.append(chunk_data['chunk_content'])

            if ids:
                self.collection.add(
                    ids=ids,
                    embeddings=embeddings,
                    metadatas=metadatas,
                    documents=documents
                )

            return len(ids)
        except Exception as e:
            raise Exception(f"添加分块到数据库失败: {e}")


def get_document_update_time(file_path: Path) -> str:
    """获取文档更新时间"""
    timestamp = file_path.stat().st_mtime
    return datetime.fromtimestamp(timestamp).strftime('%Y-%m-%d %H:%M:%S')


def process_document(
    doc_path: Path,
    category: str,
    doc_name: str,
    chunk_size: int,
    chunk_overlap: int,
    doc_update_time: str
) -> dict:
    """
    处理单个文档：提取、清洗、分块、生成embedding
    """
    try:
        # 1. 提取文本
        print(f"  正在提取文档: {doc_name}")
        raw_text = DocumentExtractor.extract_text(doc_path)

        if not raw_text:
            return {
                "success": False,
                "doc_name": doc_name,
                "error": "文档内容为空"
            }

        # 2. 清洗文本
        print(f"  正在清洗文档: {doc_name}")
        clean_text = DocumentCleaner.clean_text(raw_text)

        if not clean_text:
            return {
                "success": False,
                "doc_name": doc_name,
                "error": "清洗后文档内容为空"
            }

        # 3. 文本分块
        print(f"  正在分块: {doc_name}")
        chunks = TextChunker.chunk_text(clean_text, chunk_size, chunk_overlap)

        if not chunks:
            return {
                "success": False,
                "doc_name": doc_name,
                "error": "分块失败，没有生成任何分块"
            }

        # 4. 生成embedding
        print(f"  正在生成embedding: {doc_name} ({len(chunks)}个分块)")
        embeddings = EmbeddingGenerator.generate_embeddings(chunks)

        if len(embeddings) != len(chunks):
            return {
                "success": False,
                "doc_name": doc_name,
                "error": f"embedding生成数量不匹配: {len(embeddings)} vs {len(chunks)}"
            }

        # 5. 构建分块数据
        create_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        chunks_data = []

        for idx, (chunk, embedding) in enumerate(zip(chunks, embeddings)):
            chunk_id = f"{category}_{doc_name}_{idx:03d}"

            # 估算token数（简单估算：中文字符*1.5 + 英文单词）
            token_size = int(len(chunk) * 1.2)

            chunk_data = {
                "chunk_id": chunk_id,
                "category": category,
                "doc_name": doc_name,
                "doc_update_time": doc_update_time,
                "chunk_content": chunk,
                "chunk_token_size": token_size,
                "chunk_index": idx,
                "doc_path": str(doc_path.relative_to(PROJECT_ROOT)),
                "create_time": create_time,
                "embedding": embedding
            }
            chunks_data.append(chunk_data)

        return {
            "success": True,
            "doc_name": doc_name,
            "chunks_count": len(chunks_data),
            "chunks_data": chunks_data
        }

    except Exception as e:
        return {
            "success": False,
            "doc_name": doc_name,
            "error": str(e)
        }


def update_category(
    category: str,
    chunk_size: int,
    chunk_overlap: int,
    db: VectorDatabase
) -> dict:
    """
    更新整个分类的向量数据库
    """
    try:
        # 1. 读取分类的index.json
        category_path = RAG_LIB_PATH / category
        index_path = category_path / "index.json"

        if not category_path.exists():
            return {
                "success": False,
                "message": f"分类目录不存在: {category}"
            }

        if not index_path.exists():
            return {
                "success": False,
                "message": f"分类索引文件不存在: {index_path}"
            }

        with open(index_path, 'r', encoding='utf-8') as f:
            index_data = json.load(f)

        documents = index_data.get('documents', [])

        if not documents:
            return {
                "success": False,
                "message": f"分类 '{category}' 中没有文档"
            }

        print(f"开始更新分类 '{category}'，共 {len(documents)} 个文档")

        # 2. 删除该分类下旧的分块
        print(f"正在删除分类 '{category}' 的旧分块...")
        deleted_count = db.delete_by_category(category)
        print(f"已删除 {deleted_count} 个旧分块")

        # 3. 处理每个文档
        all_chunks_data = []
        success_count = 0
        failed_docs = []

        for doc_info in documents:
            doc_name = doc_info['doc_name']
            doc_path_str = doc_info['doc_path']

            # 构建完整路径
            if doc_path_str.startswith('rag_document_lib/'):
                doc_path = PROJECT_ROOT / 'rag' / doc_path_str
            else:
                doc_path = PROJECT_ROOT / doc_path_str

            if not doc_path.exists():
                print(f"警告: 文档不存在: {doc_path}")
                failed_docs.append({
                    "doc_name": doc_name,
                    "error": "文档文件不存在"
                })
                continue

            doc_update_time = doc_info.get('update_time', get_document_update_time(doc_path))

            # 处理文档
            result = process_document(
                doc_path,
                category,
                doc_name,
                chunk_size,
                chunk_overlap,
                doc_update_time
            )

            if result['success']:
                all_chunks_data.extend(result['chunks_data'])
                success_count += 1
                print(f"  [OK] {doc_name}: {result['chunks_count']} 个分块")
            else:
                print(f"  [FAIL] {doc_name}: {result['error']}")
                failed_docs.append({
                    "doc_name": doc_name,
                    "error": result['error']
                })

        # 4. 添加新分块到数据库
        if all_chunks_data:
            print(f"\n正在添加 {len(all_chunks_data)} 个新分块到数据库...")
            added_count = db.add_chunks(all_chunks_data)
            print(f"成功添加 {added_count} 个分块")
        else:
            print("没有可添加的分块")
            added_count = 0

        # 5. 返回结果
        return {
            "success": True,
            "message": f"分类 '{category}' 更新完成",
            "category": category,
            "documents_processed": len(documents),
            "documents_success": success_count,
            "documents_failed": len(failed_docs),
            "failed_documents": failed_docs,
            "chunks_created": added_count,
            "chunks_deleted": deleted_count,
            "chunk_size": chunk_size,
            "chunk_overlap": chunk_overlap
        }

    except Exception as e:
        return {
            "success": False,
            "message": f"更新分类失败: {str(e)}"
        }


def update_document(
    document_name: str,
    chunk_size: int,
    chunk_overlap: int,
    db: VectorDatabase
) -> dict:
    """
    更新单个文档的向量数据库
    """
    try:
        # 在所有分类中查找文档
        doc_path = None
        category = None

        # 遍历所有分类目录
        if RAG_LIB_PATH.exists():
            for category_dir in RAG_LIB_PATH.iterdir():
                if category_dir.is_dir():
                    # 检查该分类下是否有该文档
                    potential_path = category_dir / document_name
                    if potential_path.exists():
                        doc_path = potential_path
                        category = category_dir.name
                        break

        if not doc_path:
            return {
                "success": False,
                "message": f"文档不存在: {document_name}（已在所有分类中查找）"
            }

        doc_name = doc_path.name

        print(f"开始更新文档: {doc_name} (分类: {category})")

        # 1. 删除该文档的旧分块
        print(f"正在删除文档 '{doc_name}' 的旧分块...")
        deleted_count = db.delete_by_document(category, doc_name)
        print(f"已删除 {deleted_count} 个旧分块")

        # 2. 处理文档
        doc_update_time = get_document_update_time(doc_path)
        result = process_document(
            doc_path,
            category,
            doc_name,
            chunk_size,
            chunk_overlap,
            doc_update_time
        )

        if not result['success']:
            return {
                "success": False,
                "message": f"处理文档失败: {result['error']}",
                "document": doc_name
            }

        # 3. 添加新分块到数据库
        print(f"\n正在添加 {len(result['chunks_data'])} 个新分块到数据库...")
        added_count = db.add_chunks(result['chunks_data'])
        print(f"成功添加 {added_count} 个分块")

        # 4. 返回结果
        return {
            "success": True,
            "message": f"文档 '{doc_name}' 更新完成",
            "document": doc_name,
            "category": category,
            "chunks_created": added_count,
            "chunks_deleted": deleted_count,
            "chunk_size": chunk_size,
            "chunk_overlap": chunk_overlap
        }

    except Exception as e:
        return {
            "success": False,
            "message": f"更新文档失败: {str(e)}"
        }


def main():
    parser = argparse.ArgumentParser(
        description='更新向量数据库',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  # 更新整个分类
  python update_vector_db.py --category 算卦 --chunk_size 300 --chunk_overlap 40

  # 更新单个文档（只需要文档名称）
  python update_vector_db.py --document 五行.md --chunk_size 300 --chunk_overlap 40

  # 使用默认分块参数
  python update_vector_db.py --category 算卦

路径说明:
  - RAG知识库路径: rag/rag_document_lib
  - 向量数据库路径: rag/chroma_db
  - 脚本会自动拼接完整路径，无需用户输入完整路径
        """
    )

    parser.add_argument('--category', type=str, required=False,
                       help='分类名称（与--document二选一）')
    parser.add_argument('--document', type=str, required=False,
                       help='文档名称（与--category二选一，只需文档名，不需要完整路径）')
    parser.add_argument('--chunk_size', type=int, default=300,
                       help='分块字符大小（默认300）')
    parser.add_argument('--chunk_overlap', type=int, default=40,
                       help='分块重叠字符数（默认40）')

    args = parser.parse_args()

    category = args.category
    document = args.document
    chunk_size = args.chunk_size
    chunk_overlap = args.chunk_overlap

    # 验证参数
    if not category and not document:
        result = {
            "success": False,
            "message": "必须提供--category或--document参数"
        }
        print(json.dumps(result, ensure_ascii=False, indent=2))
        return

    if category and document:
        result = {
            "success": False,
            "message": "--category和--document参数不能同时使用"
        }
        print(json.dumps(result, ensure_ascii=False, indent=2))
        return

    # 验证分块参数
    if chunk_size <= 0:
        result = {
            "success": False,
            "message": "chunk_size必须大于0"
        }
        print(json.dumps(result, ensure_ascii=False, indent=2))
        return

    if chunk_overlap < 0:
        result = {
            "success": False,
            "message": "chunk_overlap不能为负数"
        }
        print(json.dumps(result, ensure_ascii=False, indent=2))
        return

    if chunk_overlap >= chunk_size:
        result = {
            "success": False,
            "message": "chunk_overlap必须小于chunk_size"
        }
        print(json.dumps(result, ensure_ascii=False, indent=2))
        return

    # 初始化向量数据库
    try:
        db = VectorDatabase(CHROMA_DB_PATH)
    except Exception as e:
        result = {
            "success": False,
            "message": f"初始化向量数据库失败: {str(e)}"
        }
        print(json.dumps(result, ensure_ascii=False, indent=2))
        return

    # 执行更新
    if category:
        result = update_category(category, chunk_size, chunk_overlap, db)
    else:
        result = update_document(document, chunk_size, chunk_overlap, db)

    # 输出结果
    print("\n" + "="*60)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    print("="*60)


if __name__ == "__main__":
    main()
